from django.shortcuts import render, get_object_or_404, redirect
from .models import Project
from .forms import ProjectForm
from django.db.models import Q, Count
from django.http import HttpResponse, JsonResponse
import csv
import openpyxl
from openpyxl.utils import get_column_letter
from django.core.paginator import Paginator
import csv
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from .models import Project
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required




# Aircraft lookup table (you can later move to DB table)
AIRCRAFT_DATA = {
    "LSP1": {"meetings": 3, "iom": 2, "sorties": 5},
    "LSP2": {"meetings": 1, "iom": 0, "sorties": 2},
    "LSP3": {"meetings": 0, "iom": 0, "sorties": 0},
    "NP1": {"meetings": 4, "iom": 1, "sorties": 6},
    "NP2": {"meetings": 2, "iom": 2, "sorties": 3},
    "PV3": {"meetings": 2, "iom": 1, "sorties": 8},
}

def project_list(request):
    aircraft = request.GET.get('aircraft','')
    status = request.GET.get('status','')
    q = request.GET.get('q','')
    page = request.GET.get('page',1)

    projects = Project.objects.all().order_by('id')
    if aircraft and aircraft != 'All':
        projects = projects.filter(aircraft__iexact=aircraft)
    if status and status != 'All':
        projects = projects.filter(status=status)
    if q:
        projects = projects.filter(Q(project_name__icontains=q) | Q(brief__icontains=q))

    aircraft_list = list(AIRCRAFT_DATA.keys())
    paginator = Paginator(projects, 20)
    page_obj = paginator.get_page(page)

    context = {'projects': page_obj, 'aircraft_list': aircraft_list, 'status_choices': ['All'] + [c[0] for c in Project._meta.get_field('status').choices],
               'selected_aircraft': aircraft or 'All', 'selected_status': status or 'All', 'q': q}
    return render(request, 'tracker/project_list.html', context)

def project_create(request):
    if request.method == 'POST':
        form = ProjectForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('tracker:list')
    else:
        form = ProjectForm()
    return render(request, 'tracker/project_form.html', {'form':form,'create':True})

def project_edit(request, pk):
    obj = get_object_or_404(Project, pk=pk)
    if request.method == 'POST':
        form = ProjectForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            return redirect('tracker:list')
    else:
        form = ProjectForm(instance=obj)
    return render(request, 'tracker/project_form.html', {'form':form,'create':False,'object':obj})

def project_delete(request, pk):
    obj = get_object_or_404(Project, pk=pk)
    if request.method == 'POST':
        obj.delete()
        return redirect('tracker:list')
    return render(request, 'tracker/project_confirm_delete.html', {'object': obj})

def export_csv(request):
    projects = Project.objects.all().order_by('id')
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=projects.csv'
    writer = csv.writer(response)
    writer.writerow(['sl.no','Project name','Brief','Status','Start','Aircraft','No.Meetings','IOM/TM','No.Sorties','Trials','End'])
    for p in projects:
        writer.writerow([p.id,p.project_name,p.brief,p.status,p.start_date,p.aircraft,p.no_meetings,p.iom_tm,p.no_sorties,p.trials,p.end_date])
    return response

def export_excel(request):
    projects = Project.objects.all().order_by('id')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Projects"
    headers = ['sl.no','Project name','Brief','Status','Start','Aircraft','No.Meetings','IOM/TM','No.Sorties','Trials','End']
    ws.append(headers)
    for p in projects:
        ws.append([p.id,p.project_name,p.brief,p.status,str(p.start_date) if p.start_date else '',p.aircraft,p.no_meetings,p.iom_tm,p.no_sorties,p.trials,str(p.end_date) if p.end_date else ''])
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length+2, 60)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=projects.xlsx'
    wb.save(response)
    return response

def stats_json(request):
    status_counts = list(Project.objects.values('status').annotate(count=Count('id')))
    aircraft_counts = list(Project.objects.values('aircraft').annotate(count=Count('id')).order_by('-count')[:20])
    return JsonResponse({'status': status_counts, 'aircraft': aircraft_counts})

def update_aircraft(request, pk, aircraft):
    obj = get_object_or_404(Project, pk=pk)
    obj.aircraft = aircraft
    data = AIRCRAFT_DATA.get(aircraft, {"meetings":0,"iom":0,"sorties":0})
    obj.no_meetings = data["meetings"]
    obj.iom_tm = data["iom"]
    obj.no_sorties = data["sorties"]
    obj.save()
    return JsonResponse({"no_meetings": obj.no_meetings, "iom_tm": obj.iom_tm, "no_sorties": obj.no_sorties})

def export_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="projects.csv"'

    writer = csv.writer(response)
    writer.writerow(['Name', 'Brief', 'Status', 'Start', 'End'])

    for p in Project.objects.all():
        writer.writerow([p.name, p.brief, p.status, p.start_date, p.end_date])

    return response

def export_excel(request):
    projects = Project.objects.all().order_by('id')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Projects"

    headers = [
        'Sl.No', 'Project Name', 'Brief', 'Status',
        'Start Date', 'Aircraft', 'No. Meetings',
        'IOM/TM', 'No. Sorties', 'Trials', 'End Date'
    ]
    ws.append(headers)

    for p in projects:
        ws.append([
            p.id,
            p.project_name,
            p.brief,
            p.status,
            p.start_date.strftime('%d-%m-%Y') if p.start_date else '',
            p.aircraft,
            p.no_meetings,
            p.iom_tm,
            p.no_sorties,
            p.trials,
            p.end_date.strftime('%d-%m-%Y') if p.end_date else '',
        ])

    # Auto column width
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 3, 50)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=projects.xlsx'

    wb.save(response)
    return response


def export_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="projects.pdf"'

    p = canvas.Canvas(response)
    y = 800

    p.setFont("Helvetica", 10)
    p.drawString(200, y, "Project Report")
    y -= 40

    for proj in Project.objects.all():
        line = f"{proj.name} | {proj.status} | {proj.start_date}"
        p.drawString(40, y, line)
        y -= 20
        if y < 50:
            p.showPage()
            y = 800

    p.showPage()
    p.save()
    return response


@login_required
@user_passes_test(lambda u: u.is_staff)
def project_delete(request, pk):
    obj = get_object_or_404(Project, pk=pk)
    if request.method == 'POST':
        obj.delete()
        return redirect('tracker:list')
    return render(request, 'tracker/project_confirm_delete.html', {'object': obj})

def login_view(request):
    if request.method == 'POST':
        user = authenticate(
            request,
            username=request.POST['username'],
            password=request.POST['password']
        )
        if user:
            login(request, user)
            return redirect('tracker:list')
        else:
            return render(request, 'tracker/login.html', {'error': 'Invalid login'})
    return render(request, 'tracker/login.html')

@login_required
def project_list(request):
    ...